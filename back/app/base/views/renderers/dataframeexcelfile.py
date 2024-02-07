from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from rest_framework import renderers

from app.base.views.renderers.utilities import (
    apply_fields_reindex,
    format_list_columns,
    get_df_from_response,
)


def initialize_workbook(df):
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    return wb


def apply_styles_and_formats(ws, fields):  # noqa: C901, PLR0912
    last_row = len(list(ws.rows))
    for index, column_key in enumerate(fields):
        # columns[column_key] = fields[column_key].get("name", column_key)
        column_letter = get_column_letter(index + 1)

        ws.column_dimensions[column_letter].width = fields[column_key].get("width", 10)

        for row in ws[column_letter + "1:" + column_letter + str(last_row)]:
            for cell in row:
                format = numbers.FORMAT_GENERAL
                column_type = fields[column_key].get("type")
                if column_type == "date":
                    format = numbers.FORMAT_DATE_DDMMYY
                if column_type == "float":
                    format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

                cell.number_format = format

    thin = Side(border_style="thin", color="000000")
    for _row in ws["1:" + str(last_row)]:
        for cell in _row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def apply_header_styles(ws):
    header_row_color = Font(name="Calibri", color="FFFFFF", bold=True)
    header_row_style = PatternFill(
        start_color="025EAA", end_color="025EAA", fill_type="solid"
    )
    for cell in ws["1"]:
        cell.font = header_row_color
        cell.fill = header_row_style
        cell.alignment = Alignment(horizontal="center")


def apply_footer_styles(ws):
    last_row = len(list(ws.rows))
    footer_color = Font(name="Calibri", color="333333", italic=True)
    total_row_style = PatternFill(
        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
    )
    for cell in ws[str(last_row)]:
        cell.font = footer_color
        cell.fill = total_row_style


class DataFrameExcelFileRenderer(renderers.BaseRenderer):
    media_type = "application/vnd.ms-excel"
    format = "xls"
    charset = None
    render_style = "binary"
    fields = {}

    def render(self, response, media_type=None, renderer_context=None):
        response_df = get_df_from_response(response)
        response_df = format_list_columns(response_df)
        # Unnecessary because openpyxl puts his own format
        # response_df = format_float_columns(response_df)
        # response_df = format_date_columns(response_df, "%d/%m/%Y")

        if renderer_context is not None:
            path_split = renderer_context["request"].get_full_path().split("/")
            renderer_context["response"][
                "Content-Disposition"
            ] = "attachment; filename={0}.xls".format(
                "{0}_{1}".format(
                    datetime.now().strftime("%Y%m%d%H%M%S"), "_".join(path_split[-2:])
                )
            )

        if self.fields:
            response_df = apply_fields_reindex(response_df, self.fields)

        wb = initialize_workbook(response_df)
        ws = wb.active
        apply_styles_and_formats(ws, self.fields)
        apply_header_styles(ws)
        apply_footer_styles(ws)

        buffer = BytesIO()
        wb.save(buffer)

        return buffer.getvalue()
